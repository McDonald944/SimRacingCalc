import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta
import pytz
import pandas as pd
import data_handling as dh
import csv
from functools import partial

def open_availability_window(master):
    """Opens the driver availability input window."""

    def create_availability_input_window(master, filepath, race_length, driver_df):
        availability_window = tk.Toplevel(master)
        try:
            icon = tk.PhotoImage(file='racing_flag_PNG.png')
            availability_window.iconphoto(False, icon)
        except:
            print("icon not found")
        availability_window.configure(bg='#3d3d3d')
        availability_window.title("Driver Availability Input")
        availability_window.geometry("450x400")  # Increased size slightly

        # Input fields
        ttk.Label(availability_window, text="Race Start Time (GMT):", font=("Arial", 12)).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        start_gmt_entry = ttk.Entry(availability_window, width=20)
        start_gmt_entry.grid(row=0, column=1, sticky="ew", padx=10, pady=5)

        ttk.Label(availability_window, text="Race Start Time (Sim Time):", font=("Arial", 12)).grid(row=1, column=0, sticky="w", padx=10, pady=5)
        start_local_entry = ttk.Entry(availability_window, width=20)
        start_local_entry.grid(row=1, column=1, sticky="ew", padx=10, pady=5)
        start_local_entry.insert(0, datetime.now().strftime("%H:%M"))  # Pre-filled with current time

        ttk.Label(availability_window, text="Green Flag Offset (minutes):", font=("Arial", 12)).grid(row=2, column=0, sticky="w", padx=10, pady=5)
        offset_entry = ttk.Entry(availability_window, width=20)
        offset_entry.insert(0, "30")
        offset_entry.grid(row=2, column=1, sticky="ew", padx=10, pady=5)

        # Time Block Frame
        time_block_label = ttk.Label(availability_window, text="Time Block (minutes):", font=("Arial", 12))
        time_block_label.grid(row=3, column=0, sticky="nw", padx=10, pady=5)
        time_block_frame = ttk.Frame(availability_window)
        time_block_frame.grid(row=3, column=1, sticky="nsew", padx=10, pady=5)
        availability_window.rowconfigure(3, weight=1)
        time_block = tk.IntVar(value=15)

        time_block_values = [15, 30, 45, 60]
        for i, value in enumerate(time_block_values):
            ttk.Radiobutton(time_block_frame, text=f"{value} Minutes", variable=time_block, value=value).grid(row=i, column=0, sticky="w", pady=2)

        def create_availability(availability_window):
            offset = offset_entry.get()
            start_gmt_str = start_gmt_entry.get()
            start_local_str = start_local_entry.get() if start_local_entry.get() else None

            generate_availability_sheet(filepath, start_gmt_str, start_local_str, int(offset), time_block.get(), race_length, master, availability_window)  # pass the filepath

        # Center the window
        availability_window.update_idletasks()
        width = availability_window.winfo_width()
        height = availability_window.winfo_height()
        x = (master.winfo_screenwidth() // 2) - (width // 2)
        y = (master.winfo_screenheight() // 2) - (height // 2)
        availability_window.geometry("+{}+{}".format(x, y))

        create_button_command = partial(create_availability, availability_window)  # Create the partial function
        create_button = ttk.Button(availability_window, text="Create Availability Sheet", command=create_button_command) # Assign the partial function to command
        create_button.grid(row=4, column=0, columnspan=2, pady=(10, 20), sticky="ew")

    filepath = dh.open_roster_file()
    if filepath:
        result = dh.check_roster_data(filepath)
        if result:
            race_length, driver_df = result
            if driver_df is not None and race_length is not None:
                create_availability_input_window(master, filepath, race_length, driver_df)
            else:
                messagebox.showerror("Error", "Could not read data from roster file.")
        else:
            messagebox.showerror("Error", "Could not open roster file.")




def generate_availability_sheet(filepath, start_gmt_str, start_local_str, offset_str, time_block, race_length_str, master, availability_window): #take filepath as a parameter
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Driver Availability"
        data = []
        with open(filepath, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if row:
                    data.append(row)
        print("Raw Data from CSV:", data)  # print the raw data

        if len(data) < 4:
            messagebox.showerror("Error", "Roster file must contain event and driver data.")
            return None

        # Correctly create event_data dictionary
        event_data = {}
        for i in range(len(data[0])):  # Iterate through the columns
            if i < len(data[1]):  # check to make sure the index exists in the second row
                event_data[data[0][i]] = data[1][i]
            else:
                event_data[data[0][i]] = ""  # if it doesnt exist, make it an empty string
        print("Event Data:", event_data)

        if "Race Length" not in event_data:
            messagebox.showerror("Error", "Roster file must contain 'Race Length' data.")
            return None
        try:
            race_length_parts = list(map(int, event_data["Race Length"].split(':')))
            if len(race_length_parts) == 2:
                race_length = timedelta(hours=race_length_parts[0], minutes=race_length_parts[1])
            elif len(race_length_parts) == 3:
                race_length = timedelta(days=race_length_parts[0], hours=race_length_parts[1],
                                        minutes=race_length_parts[2])
            else:
                raise ValueError("Invalid race length format. Use HH:MM or D:HH:MM")
            print("Race Length:", race_length)

        except ValueError as e:
            messagebox.showerror("ValueError", f"Invalid race length format: {e}")
            return

        driver_data = data[2:]  # start at the fourth line
        print("Driver Data:", driver_data)
        if not driver_data or not driver_data[0]:
            messagebox.showerror("Error", "Roster file must contain driver data.")
            return

        roster_df = pd.DataFrame(driver_data[1:], columns=driver_data[0])

        start_gmt = datetime.strptime(start_gmt_str, "%H:%M").replace(tzinfo=pytz.utc)
        offset = int(offset_str)

        race_start_gmt = start_gmt + timedelta(minutes=offset)
        race_end_gmt = race_start_gmt + race_length

        # Write Time Headers
        sheet.cell(row=1, column=1).value = "Race Time"
        sheet.cell(row=1, column=2).value = "GMT"
        sim_time_col = 3  # Column for Sim Time (if needed)
        if start_local_str:
            sheet.cell(row=1, column=sim_time_col).value = "Sim Time"

        # Driver Headers
        driver_start_col = sim_time_col + 1  # Start column for driver data
        driver_col_index = driver_start_col
        for driver_name in roster_df["Driver Name"]:  # iterate through the driver names
            sheet.cell(row=1, column=driver_col_index).value = f"{driver_name} Local"
            sheet.cell(row=1, column=driver_col_index + 1).value = f"{driver_name} Availability"
            driver_col_index += 2

        current_time_gmt = race_start_gmt
        row_num = 2
        while current_time_gmt <= race_end_gmt:
            sheet.cell(row=row_num, column=1).value = (current_time_gmt - race_start_gmt).seconds // 3600
            sheet.cell(row=row_num, column=2).value = current_time_gmt.strftime("%H:%M")
            if start_local_str:
                local_time = (datetime.combine(datetime.today(), datetime.strptime(start_local_str, "%H:%M").time()) + (
                            current_time_gmt - race_start_gmt)).time()
                sheet.cell(row=row_num, column=sim_time_col).value = local_time.strftime("%H:%M")

            # Driver Local Time and Availability
            driver_data_row = 0
            driver_col_index = driver_start_col
            for _, driver_row_data in roster_df.iterrows():
                timezone_str = driver_row_data["Timezone"]
                try:
                    driver_timezone = pytz.timezone(timezone_str)
                except pytz.UnknownTimeZoneError:
                    messagebox.showerror("Timezone Error",
                                         f"Unknown timezone: {timezone_str} for driver {driver_row_data['Driver Name']}")
                    return
                driver_gmt = current_time_gmt.astimezone(driver_timezone)
                sheet.cell(row=row_num, column=driver_col_index).value = driver_gmt.strftime("%H:%M")
                sheet.cell(row=row_num,
                           column=driver_col_index + 1).value = "Available\nMonitor\nTentative\nSleep\nBlocked"
                driver_col_index += 2
                driver_data_row += 1

            current_time_gmt += timedelta(minutes=time_block)
            row_num += 1

        # Apply Formatting
        bold_font = Font(bold=True)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border
                if row[0].row == 1:  # format the header row
                    cell.font = bold_font
                if cell.column == 4:  # format availability column
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value]  # check if the cell has a value
            if column:  # check if the column list is not empty
                try:
                    max_length = max(len(str(cell.value)) for cell in column)
                except ValueError:  # if there is a value error, set the max length to 0
                    max_length = 0
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        availability_options = ["Available", "Monitor", "Tentative", "Sleep", "Blocked"]
        for row in range(2, row_num):
            for driver_index in range(0, len(roster_df)):
                availability_col = driver_start_col + (driver_index * 2) + 1
                cell = sheet.cell(row=row, column=availability_col)
                dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"{}"'.format(
                    ",".join(availability_options)))
                sheet.add_data_validation(dv)
                dv.add(cell)
                cell.value = "Tentative"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Driver Availability Sheet",
            initialfile="driver_availability.xlsx",
            parent=master
        )

        if filepath:
            try:
                wb.save(filepath)
                messagebox.showinfo("Success", f"Driver Availability sheet saved to {filepath}!")
                availability_window.destroy()
            except OSError as e:
                messagebox.showerror("Save Error", f"Error saving file: {e}")
            except Exception as e:
                messagebox.showerror("Save Error", f"An unexpected error occurred during save: {e}")
        else:
            messagebox.showinfo("Info", "Save operation cancelled.")

    except ValueError as e:
        messagebox.showerror("ValueError", f"A value error has occurred: {e}")
    except pytz.UnknownTimeZoneError as e:
        messagebox.showerror("Timezone Error", f"Invalid Timezone: {e}")
    except Exception as e:
        print(f"Full Exception: {e}")  # Print the full exception information for debugging
        messagebox.showerror("Error", f"Error generating Excel sheet: {str(e)}")